from openpyxl import load_workbook
import sqlite3 as sql
from flask import Flask, request, render_template, redirect, url_for,jsonify,flash
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy.ext.automap import automap_base
from sqlalchemy.orm import Session
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

app = Flask(__name__)
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///revenuenew.sqlite3'
app.config['SECRET_KEY'] = "random string"
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db = SQLAlchemy(app)

class revenuenew(db.Model):
   id = db.Column('id', db.Integer, primary_key=True)
   department = db.Column(db.String(30))
   region = db.Column(db.String(30))
   service_line = db.Column(db.String(50))
   actual_mtd = db.Column(db.Integer)
   budget_mtd = db.Column(db.Integer)
   prior_mtd = db.Column(db.Integer)
   actual_ytd = db.Column(db.Integer)
   budget_ytd = db.Column(db.Integer)
   prior_ytd = db.Column(db.Integer)


   def __init__(self, id, department, region, service_line,actual_mtd,budget_mtd,prior_mtd,actual_ytd,budget_ytd,prior_ytd):
    self.id = id
    self.department = department
    self.region = region
    self.service_line = service_line
    self.actual_mtd = actual_mtd
    self.budget_mtd = budget_mtd
    self.prior_mtd = prior_mtd
    self.actual_ytd = actual_ytd
    self.budget_ytd = budget_ytd
    self.prior_ytd = prior_ytd

db.create_all()


@app.route('/')
def index():
    return """<title>Upload new File</title>
    <h1>Upload new File</h1>
    <form action="/uploader" method=post enctype=multipart/form-data>
      <p><input type=file name=file>
         <input type=submit value=Upload>
    </form>
    <form action="/show" method=post enctype=multipart/form-data>
         <input type=submit value=data>
    </form>
    """

@app.route('/uploader', methods = ['GET', 'POST'])
def upload():
    if request.method == 'POST':
        f = request.files['file']
        f.save(f.filename)
        return process(f.filename)

def process(filename):


    wb = load_workbook(filename)
    sheet = wb.active

    wb.save(filename=r"C:\Users\Ashish.Gupta\Desktop\hello_world.xlsx")
    sheetnames=wb.sheetnames
    rows = sheet.max_row
    columns = sheet.max_column

    if request.method == 'POST':
            id = sheet.cell(row=5, column=1).value
            department = sheet.cell(row=5, column=2).value
            region = sheet.cell(row=5, column=3).value
            service_line = sheet.cell(row=5, column=4).value
            actual_mtd = sheet.cell(row=5, column=5).value
            budget_mtd = sheet.cell(row=5, column=6).value
            prior_mtd = sheet.cell(row=5, column=7).value
            actual_ytd = sheet.cell(row=5, column=8).value
            budget_ytd = sheet.cell(row=5, column=9).value
            prior_ytd = sheet.cell(row=5, column=10).value

            data = revenuenew(id,department,region,service_line,actual_mtd,budget_mtd,prior_mtd,actual_ytd,budget_ytd,prior_ytd)
            db.session.add(data)
            db.session.commit()
            msg = "Record successfully added"

    return render_template('result.html')


@app.route('/show',methods=['POST', 'GET'])
def show():
   return render_template('show.html', revenuenew=revenuenew.query.all())


@app.route('/department',methods=['POST', 'GET'])
def show():
   result = revenuenew.query.filter_by(department='C&I').all()
   return jsonify(result)


if __name__ == "__main__":
    app.run()