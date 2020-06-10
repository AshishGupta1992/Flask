from openpyxl import load_workbook
from flask import Flask, request, render_template, redirect, url_for,jsonify


app = Flask(__name__)

@app.route('/')
def index():
    return """<title>Upload new File</title>
    <h1>Upload new File</h1>
    <form action="/uploader" method=post enctype=multipart/form-data>
      <p><input type=file name=file>
         <input type=submit value=Upload>
    </form>"""

@app.route('/uploader', methods = ['GET', 'POST'])
def upload():
    if request.method == 'POST':
        f = request.files['file']
        f.save(f.filename)
        return process(f.filename)

def process(filename):


    wb = load_workbook(filename)
    sheet = wb.active

    wb.save(filename=r"C:\Users\Ashish.Gupta\Desktop\hello_world.xlsx")
    sheetnames=wb.sheetnames
    rows = sheet.max_row
    columns = sheet.max_column
    return jsonify(columns)


if __name__ == "__main__":
    app.run()