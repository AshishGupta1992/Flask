from openpyxl import load_workbook
import sqlite3 as sql
from flask import Flask, request, render_template, redirect, url_for,jsonify


app = Flask(__name__)

@app.route('/')
def index():
    return """<title>Upload new File</title>
    <h1>Upload new File</h1>
    <form action="/uploader" method=post enctype=multipart/form-data>
      <p><input type=file name=file>
         <input type=submit value=Upload>
    </form>
    <form action="/list" method=post enctype=multipart/form-data>
         <input type=submit value=data>
    </form>
    """

@app.route('/uploader', methods = ['GET', 'POST'])
def upload():
    if request.method == 'POST':
        f = request.files['file']
        f.save(f.filename)
        return process(f.filename)

def process(filename):


    wb = load_workbook(filename)
    sheet = wb.active

    wb.save(filename=r"C:\Users\Ashish.Gupta\Desktop\hello_world.xlsx")
    sheetnames=wb.sheetnames
    rows = sheet.max_row
    columns = sheet.max_column

    if request.method == 'POST':
        try:
            department = sheet.cell(row=2, column=1).value
            region = sheet.cell(row=2, column=2).value
            service_line = sheet.cell(row=2, column=3).value
            actual_mtd = sheet.cell(row=2, column=4).value
            actual_mtd = sheet.cell(row=2, column=5).value
            actual_mtd = sheet.cell(row=2, column=6).value
            budget_mtd = sheet.cell(row=2, column=7).value
            prior_mtd = sheet.cell(row=2, column=8).value
            actual_ytd = sheet.cell(row=2, column=9).value
            budget_ytd = sheet.cell(row=2, column=10).value
            prior_ytd = sheet.cell(row=2, column=11).value

            with sql.connect("revenue.db") as con:
                cur = con.cursor()

                cur.execute("INSERT INTO pl (department, region, service_line, actual_mtd, budget_mtd, prior_mtd, actual_ytd, budget_ytd, prior_ytd) VALUES(?, ?, ?, ?,?, ?, ?, ?,?)",(department, region, service_line, actual_mtd, budget_mtd, prior_mtd, actual_ytd, budget_ytd, prior_ytd) )

                con.commit()
                msg = "Record successfully added"
        except:
            con.rollback()
            msg = "error in insert operation"

        finally:
            return render_template("result.html", msg=msg)
            con.close()


    return jsonify(columns)

@app.route('/list',methods=['POST', 'GET'])
def list():
    con = sql.connect("revenue.db")
    con.row_factory = sql.Row

    cur = con.cursor()
    cur.execute("select * from pl")

    rows = cur.fetchall();
    return render_template("list.html", rows=rows)


if __name__ == "__main__":
    app.run()